from pathlib import Path
from typing import Optional, List
import pandas as pd
from openpyxl import load_workbook


def find_first_data_row(file_path: Path) -> int:
    """
    Функция для определения первой строки с данными в Excel файле.

    Args:
        file_path (Path): Путь к Excel файлу.

    Returns:
        int: Индекс первой строки с данными.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(row) and not all(cell is None for cell in row):
            return i
    return 0


def read_excel_values(file_path: Path) -> pd.DataFrame:
    """
    Функция для считывания всех значений из Excel файла.

    Args:
        file_path (Path): Путь к Excel файлу.

    Returns:
        pd.DataFrame: DataFrame со всеми значениями из Excel файла.
    """
    start_row = find_first_data_row(file_path)
    df = pd.read_excel(file_path, skiprows=start_row)
    if not df.empty:
        # Удаление колонок с именем "Unnamed"
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    return df


def mask_columns(df: pd.DataFrame, columns_to_mask: List[int]) -> pd.DataFrame:
    """
    Функция для замены данных в указанных столбцах на маскированные значения.

    Args:
        df (pd.DataFrame): DataFrame с данными.
        columns_to_mask (List[int]): Список индексов столбцов, данные в которых нужно заменить.

    Returns:
        pd.DataFrame: DataFrame с замаскированными данными.
    """
    for column_idx in columns_to_mask:
        if column_idx < len(df.columns):
            column = df.columns[column_idx]
            df[column] = df[column].apply(lambda x: '*' * len(str(x)) if pd.notnull(x) else x)
    return df


def save_censored_file(df: pd.DataFrame, original_file_path: Path) -> Path:
    """
    Функция для сохранения замаскированных данных в новый файл с постфиксом _censored.

    Args:
        df (pd.DataFrame): DataFrame с замаскированными данными.
        original_file_path (Path): Путь к оригинальному файлу Excel.

    Returns:
        Path: Путь к новому файлу с замаскированными данными.
    """
    censored_file_path = original_file_path.with_name(f"{original_file_path.stem}_censored{original_file_path.suffix}")
    df.to_excel(censored_file_path, index=False)
    return censored_file_path
